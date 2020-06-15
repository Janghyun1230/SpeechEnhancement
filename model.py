import matplotlib as mpl
mpl.use('Agg')
import matplotlib.pyplot as plt
import tensorflow as tf
import numpy as np
from glob import glob
import os
from shutil import copyfile
from time import sleep, strftime, localtime
from Loader import Loader
from configuration import config
from Utils import *
from layers import *
from net import FrontEnd, Unet, LSTM, BiLSTM, Dilated, BackEnd, Confidence
import librosa
import time
import openpyxl
from scipy import signal
from loss import loss_fn, loss_alpha
from build_graph import build_graph
import nsml

tf.reset_default_graph()
np.set_printoptions(precision=4)
                       

class MDPHD(object):
    def __init__(self, name):
        self.name = name
        self.loader = Loader([config.train_root], config.batch_size)

    def build(self, reuse=False):
        self.reuse = reuse
        with tf.variable_scope(self.name, reuse=self.reuse):
            # train set
            self.mix = self.loader.mix_queue # mixture batch
            self.source = self.loader.source_queue # source batch
            self.other = self.mix - self.source
            
            # test samples (placeholder)
            self.mix_test = tf.placeholder(tf.float32, shape=(1,16384), name="mixture_sample")
            self.source_test = tf.placeholder(tf.float32, shape=(1,16384), name="source_sample")

            # define network
            self.FE = FrontEnd()
            if config.network == "unet":
                self.SEP = Unet(model=config.model)
            elif config.network == "lstm":
                self.SEP = LSTM(model=config.model)
            elif config.network == "bilstm":
                self.SEP = BiLSTM(model=config.model)
            elif config.network == "dilated":
                self.SEP = Dilated(model=config.model)
            else:
                print("No model chosen")
                raise ValueError
            self.BE = BackEnd()
            functions = [self.FE, self.SEP, self.BE]

            if config.network2 == "unet":
                self.SEP2 = Unet(model=config.model2)
                self.CON = Confidence(model=config.confidence)
                functions = functions + [self.SEP2, self.CON]
            elif config.network2 == "lstm":
                self.SEP2 = LSTM(model=config.model2)
                self.CON = Confidence(model=config.confidence)
                functions = functions + [self.SEP2, self.CON]
            elif config.network2 == "bilstm":
                self.SEP2 = BiLSTM(model=config.model2)
                self.CON = Confidence(model=config.confidence)
                functions = functions + [self.SEP2, self.CON]
            elif config.network2 == "dilated":
                self.SEP2 = Dilated(model=config.model2)
                self.CON = Confidence(model=config.confidence)
                functions = functions + [self.SEP2, self.CON]
            else:
                print("No model chosen")

            input = {"mix": self.mix, "source": self.source}
            input_test = {"mix": self.mix_test, "source": self.source_test}

            # draw graph
            self.graph = build_graph()
            self.graph_test = build_graph(train=False)

            self.graph(input, functions)
            self.update_ops = tf.get_collection(tf.GraphKeys.UPDATE_OPS)
            self.graph_test(input_test, functions)

            # variable count
            self.variable_size = np.sum(
                np.array([np.prod(np.array(v.get_shape().as_list())) for v in tf.trainable_variables()]))
            print("\n total varialbes : {}".format(self.variable_size), "\n")
            for v in tf.trainable_variables():
                print(v, "{}".format(np.prod(np.array(v.get_shape().as_list()))))

            # Define loss and summarize
            # self.loss_pre = (loss_fn(self.graph.masked_spec1, self.graph.estimated1, self.graph.source_spec,
            #                     self.source, self.other, loss_type=config.loss) + \
            #                 loss_fn(self.graph.masked_spec2, self.graph.estimated2, self.graph.source_spec,
            #                         self.source, self.other, loss_type=config.loss))/2

            if config.hybrid:
                if config.loss_seq == 'mid':
                    self.loss = loss_fn(self.graph.masked_spec1_mid, self.graph.estimated1_mid, self.graph.source_spec,
                                        self.source, self.other, loss_type=config.loss) + \
                                loss_fn(self.graph.masked_spec2_mid, self.graph.estimated2_mid, self.graph.source_spec,
                                        self.source, self.other, loss_type=config.loss)
                elif config.loss_seq == 'end':
                    self.loss = loss_fn(self.graph.masked_spec1, self.graph.estimated1, self.graph.source_spec,
                                        self.source, self.other, loss_type=config.loss) + \
                                loss_fn(self.graph.masked_spec2, self.graph.estimated2, self.graph.source_spec,
                                        self.source, self.other, loss_type=config.loss)
                elif config.loss_seq == 'both':
                    self.loss = (loss_fn(self.graph.masked_spec1, self.graph.estimated1, self.graph.source_spec,
                                        self.source, self.other, loss_type=config.loss) +
                                loss_fn(self.graph.masked_spec2, self.graph.estimated2, self.graph.source_spec,
                                        self.source, self.other, loss_type=config.loss) +
                                loss_fn(self.graph.masked_spec1_mid, self.graph.estimated1_mid, self.graph.source_spec,
                                        self.source, self.other, loss_type=config.loss) +
                                loss_fn(self.graph.masked_spec2_mid, self.graph.estimated2_mid, self.graph.source_spec,
                                        self.source, self.other, loss_type=config.loss))/2
                elif config.loss_seq == 'two':
                    self.loss_pre = loss_fn(self.graph.masked_spec1_mid, self.graph.estimated1_mid, self.graph.source_spec,
                                        self.source, self.other, loss_type=config.loss) + \
                                loss_fn(self.graph.masked_spec2_mid, self.graph.estimated2_mid, self.graph.source_spec,
                                        self.source, self.other, loss_type=config.loss)

                    self.loss = (loss_fn(self.graph.masked_spec1, self.graph.estimated1, self.graph.source_spec,
                                         self.source, self.other, loss_type=config.loss) +
                                 loss_fn(self.graph.masked_spec2, self.graph.estimated2, self.graph.source_spec,
                                         self.source, self.other, loss_type=config.loss) +
                                 loss_fn(self.graph.masked_spec1_mid, self.graph.estimated1_mid, self.graph.source_spec,
                                         self.source, self.other, loss_type=config.loss) +
                                 loss_fn(self.graph.masked_spec2_mid, self.graph.estimated2_mid, self.graph.source_spec,
                                         self.source, self.other, loss_type=config.loss)) / 2
                elif config.loss_seq == 'first':
                    self.loss_pre = loss_fn(self.graph.masked_spec1_mid, self.graph.estimated1_mid, self.graph.source_spec,
                                        self.source, self.other, loss_type=config.loss)

                    self.loss = loss_fn(self.graph.masked_spec1, self.graph.estimated1, self.graph.source_spec,
                                         self.source, self.other, loss_type=config.loss) + \
                                 loss_fn(self.graph.masked_spec1_mid, self.graph.estimated1_mid, self.graph.source_spec,
                                         self.source, self.other, loss_type=config.loss)
                    self.graph.estimated = self.graph.estimated1
                    self.graph_test.estimated = self.graph_test.estimated1
                elif config.loss_seq == 'second':
                    self.loss_pre = loss_fn(self.graph.masked_spec2_mid, self.graph.estimated2_mid, self.graph.source_spec,
                                        self.source, self.other, loss_type=config.loss)

                    self.loss = loss_fn(self.graph.masked_spec2, self.graph.estimated2, self.graph.source_spec,
                                         self.source, self.other, loss_type=config.loss) + \
                                 loss_fn(self.graph.masked_spec2_mid, self.graph.estimated2_mid, self.graph.source_spec,
                                         self.source, self.other, loss_type=config.loss)
                    self.graph.estimated = self.graph.estimated2
                    self.graph_test.estimated = self.graph_test.estimated2
                else:
                    raise AssertionError("wrong config.loss_seq !!")
            else:
                self.loss = loss_fn(self.graph.masked_spec, self.graph.estimated, self.graph.source_spec,
                                    self.source, self.other, loss_type=config.loss)

            self.loss_summary = tf.summary.scalar("loss", self.loss)
            self.alpha_summary = tf.summary.scalar("alpha", tf.reduce_mean(self.graph.alpha))
            self.summary = tf.summary.merge_all()
            print("\n loss type : %s \n" % config.loss)

    def train(self):
        with tf.control_dependencies(self.update_ops):
            step = tf.Variable(0, trainable=False)
            lr_ = config.lr
            lr = tf.placeholder(tf.float32)
            optimizer = tf.train.AdamOptimizer(lr, beta1=config.Beta1, beta2=config.Beta2)
            if (config.loss_seq in ('two', 'first', 'second')) & (config.hybrid == True):
                self.trainer_pre = optimizer.minimize(
                    self.loss_pre, var_list=tf.trainable_variables(), global_step=step)
            self.trainer = optimizer.minimize(
                self.loss, var_list=tf.trainable_variables(), global_step=step)

        NUM_THREADS = 3
        configuration = tf.ConfigProto(inter_op_parallelism_threads=NUM_THREADS, \
                                       intra_op_parallelism_threads=NUM_THREADS, \
                                       allow_soft_placement=True, \
                                       device_count={'CPU': 3}, \
                                       )

        with tf.Session(config=configuration) as sess:

            path = os.path.join(os.getcwd(), 'Results', 'Spectrograms')
            os.makedirs(path, exist_ok=True)
            path = os.path.join(os.getcwd(), 'Results', 'Test')
            os.makedirs(path, exist_ok=True)
            path = os.path.join(os.getcwd(), 'Results', 'Simulation')
            os.makedirs(path, exist_ok=True)
            path = os.path.join(os.getcwd(), 'Results', 'CheckPoint')
            os.makedirs(path, exist_ok=True)
            path = os.path.join(os.getcwd(), 'Results', 'Logs_Training')
            os.makedirs(path, exist_ok=True)

            tf.global_variables_initializer().run()

            coord = tf.train.Coordinator()
            threads = tf.train.start_queue_runners(sess, coord=coord)

            saver = tf.train.Saver(max_to_keep=10)
            writer = tf.summary.FileWriter('./Results/Logs_Training', sess.graph)

            """
            Restore
                Restore checkpoint only when the user specify the 'config.restore' to be True
            """
            if config.restore:
                ckpt = tf.train.get_checkpoint_state(checkpoint_dir="./Results/CheckPoint")
                try:
                    if ckpt and ckpt.model_checkpoint_path:
                        print("check point path : ", ckpt.model_checkpoint_path)
                        saver.restore(sess, ckpt.model_checkpoint_path)
                        print('Restored!')
                except AttributeError:
                    print("No checkpoint")
            else:
                print("No model restoration")

            training_loss = 0
            training_alpha_sum = 0
            training_alpha_list = []
            f = open(os.path.join(os.getcwd(), 'Results', 'Logs_Training', 'training_loss.txt'), 'a+')
            f.write("Variable Size : {}\n".format(self.variable_size))
            f.close()

            for iter in range(config.iter + 1):
                if coord.should_stop():
                    print("break data pipeline")
                    break

                if (config.loss_seq in ('two', 'first', 'second')) & (config.hybrid == True):
                    if iter <= config.loss_thereshold:
                        _, loss_, alpha_ = sess.run([self.trainer_pre, self.loss_pre, self.graph.alpha], feed_dict={lr: lr_})
                    else:
                        _, loss_, alpha_ = sess.run([self.trainer, self.loss, self.graph.alpha], feed_dict={lr: lr_})
                else:
                    _, loss_, alpha_ = sess.run([self.trainer, self.loss, self.graph.alpha], feed_dict={lr: lr_})

                training_loss += loss_
                training_alpha_sum += np.mean(alpha_)
                training_alpha_list.append(np.squeeze(alpha_))

                if iter % 10 == 0:
                    print("iter : {} (loss : {:.3f}, alpha : {:.3f}/ {:.3f})".format(iter, loss_, np.mean(alpha_), np.std(alpha_)))
                    summary_ = sess.run(self.summary)
                    writer.add_summary(summary_, iter)

                # print spectrograms
                if iter % 5000 == 0:
                    spec_mix_, spec_tgt_, spec_est_ = sess.run(
                        [self.graph.input_spec, self.graph.source_spec, self.graph.masked_spec])

                    mag_mix_ = np.sqrt(np.square(spec_mix_[:, :, :, 0]) + np.square(spec_mix_[:, :, :, 1]))
                    mag_tgt_ = np.sqrt(np.square(spec_tgt_[:, :, :, 0]) + np.square(spec_tgt_[:, :, :, 1]))
                    mag_est_ = np.sqrt(np.square(spec_est_[:, :, :, 0]) + np.square(spec_est_[:, :, :, 1]))

                    mag_image1 = np.log10(np.concatenate((mag_mix_[0], mag_mix_[1], mag_mix_[2]), axis=1) + 1e-02)
                    mag_image2 = np.log10(np.concatenate((mag_tgt_[0], mag_tgt_[1], mag_tgt_[2]), axis=1) + 1e-02)
                    mag_image3 = np.log10(np.concatenate((mag_est_[0], mag_est_[1], mag_est_[2]), axis=1) + 1e-02)

                    mag_image4 = np.log10(
                        np.concatenate((mag_mix_[-3], mag_mix_[-2], mag_mix_[-1]), axis=1) + 1e-02)
                    mag_image5 = np.log10(
                        np.concatenate((mag_tgt_[-3], mag_tgt_[-2], mag_tgt_[-1]), axis=1) + 1e-02)
                    mag_image6 = np.log10(
                        np.concatenate((mag_est_[-3], mag_est_[-2], mag_est_[-1]), axis=1) + 1e-02)

                    fig = plt.figure(figsize=(15, 15))
                    fig.suptitle('Spectrograms', fontsize=20, family='serif')

                    font = {'family': 'serif',
                            'color': 'darkred',
                            'weight': 'normal',
                            'size': 15,
                            }

                    ax1 = fig.add_subplot(1, 3, 1)
                    ax1.set_title('Mixture', fontdict=font)
                    ax1.imshow(mag_image1, interpolation='nearest', aspect='auto', cmap='jet')
                    ax1.xaxis.set_tick_params(labelsize=15)
                    ax1.yaxis.set_tick_params(labelsize=15)
                    plt.gca().invert_yaxis()

                    ax2 = fig.add_subplot(1, 3, 2)
                    ax2.set_title('True Vocal', fontdict=font)
                    ax2.imshow(mag_image2, interpolation='nearest', aspect='auto', cmap='jet')
                    ax2.xaxis.set_tick_params(labelsize=15)
                    ax2.yaxis.set_tick_params(labelsize=15)
                    plt.gca().invert_yaxis()

                    ax3 = fig.add_subplot(1, 3, 3)
                    ax3.set_title('Estimated Vocal', fontdict=font)
                    ax3.imshow(mag_image3, interpolation='nearest', aspect='auto', cmap='jet')
                    ax3.xaxis.set_tick_params(labelsize=15)
                    ax3.yaxis.set_tick_params(labelsize=15)
                    plt.gca().invert_yaxis()

                    plt.tight_layout(pad=0.4, w_pad=1, h_pad=1.0)
                    plt.subplots_adjust(top=0.85)
                    fig.savefig("./Results/Spectrograms/Spectrogram{}.png".format(iter))
                    plt.close(fig)

                    fig = plt.figure(figsize=(15, 15))
                    fig.suptitle('NoiseSpectrograms', fontsize=20, family='serif')

                    font = {'family': 'serif',
                            'color': 'darkred',
                            'weight': 'normal',
                            'size': 15,
                            }

                    ax1 = fig.add_subplot(1, 3, 1)
                    ax1.set_title('Noise', fontdict=font)
                    ax1.imshow(mag_image4, interpolation='nearest', aspect='auto', cmap='jet')
                    ax1.xaxis.set_tick_params(labelsize=15)
                    ax1.yaxis.set_tick_params(labelsize=15)
                    plt.gca().invert_yaxis()

                    ax2 = fig.add_subplot(1, 3, 2)
                    ax2.set_title('True Zeros', fontdict=font)
                    ax2.imshow(mag_image5, interpolation='nearest', aspect='auto', cmap='jet')
                    ax2.xaxis.set_tick_params(labelsize=15)
                    ax2.yaxis.set_tick_params(labelsize=15)
                    plt.gca().invert_yaxis()

                    ax3 = fig.add_subplot(1, 3, 3)
                    ax3.set_title('Estimated Zeros', fontdict=font)
                    ax3.imshow(mag_image6, interpolation='nearest', aspect='auto', cmap='jet')
                    ax3.xaxis.set_tick_params(labelsize=15)
                    ax3.yaxis.set_tick_params(labelsize=15)
                    plt.gca().invert_yaxis()

                    plt.tight_layout(pad=0.4, w_pad=1, h_pad=1.0)
                    plt.subplots_adjust(top=0.85)
                    fig.savefig(
                        "./Results/Spectrograms/Noise_input_Spectrogram{}.png".format(iter))
                    plt.close(fig)

                if (iter) % 10000 == 0:
                    training_alpha_std = np.std(np.array(training_alpha_list))
                    f = open(os.path.join(os.getcwd(), 'Results', 'Logs_Training', 'training_loss.txt'), 'a+')
                    f.write("average {} loss : {:.4f} (iter {} ~ {})\n".format(config.loss, training_loss / 10000,
                                                                               iter - 9999, iter + 1))
                    f.close()
                    f = open(os.path.join(os.getcwd(), 'Results', 'Logs_Training', 'training_alpha.txt'), 'a+')
                    f.write("average alpha : {:.4f} / std : {:.4f} (iter {} ~ {})\n".format(training_alpha_sum / 10000,
                                                                                            training_alpha_std ,iter - 9999, iter + 1))
                    f.close()

                    nsml.report(summary=True,
                                iter=iter,
                                train__loss=float(training_loss / 10000),
                                alpha=float(training_alpha_sum / 10000),
                                alpha_std=float(training_alpha_std),
                                lr=float(lr_)*10000
                                )
                    training_loss = 0
                    training_alpha_sum = 0
                    training_alpha_list = []

                if (iter+1) % config.lr_decay_iter == 0:
                    if config.lr_decay:
                        lr_ = lr_/config.lr_decay_rate
                        print("learning rate is decayed : {}".format(lr_))

                # test songs
                if (iter + 1) % config.model_save_iter == 0:
                    moving_rate = config.moving_rate
                    hop = int(config.audio_size / moving_rate)  # This has to be an interger without any rounding
                    mask = (signal.get_window('hanning', config.audio_size)) / (moving_rate / 2)
                    # mask = (1 / moving_rate)

                    for folder, root in [('Test', config.test_sample_root), ('Simulation', config.simul_sample_root)]:
                        test_sample_path = glob(os.path.join(root, "noisy/*.wav"))
                        print("{} audios for test".format(len(test_sample_path)))

                        SNR_list = []
                        infer_time = []

                        wb = openpyxl.Workbook()
                        ws = wb.active

                        for k, song_path in enumerate(sorted(test_sample_path)):
                            print("\n %d th song" % k)

                            # Load song
                            name = os.path.basename(song_path)
                            print("Loading {}...".format(name))
                            noisy, _ = librosa.core.load(song_path, sr=16000)
                            clean, _ = librosa.core.load(os.path.join(root, "spk", name[:8]+".wav"),
                                                         sr=16000)
                            noise = noisy - clean

                            print("song shape : ", noisy.shape)

                            song_length = noisy.shape[0]
                            num_frame = int(song_length / config.audio_size) + 1

                            # pad zeros to song to make the length of it be the multiple of config.audio_size
                            noisy_pad = np.lib.pad(noisy, (
                                config.audio_size, num_frame * config.audio_size - song_length + config.audio_size),
                                                   'constant')
                            clean_pad = np.lib.pad(clean, (
                                config.audio_size, num_frame * config.audio_size - song_length + config.audio_size),
                                                   'constant')
                            padded_song_shape = noisy_pad.shape[0]

                            # Slice mixture and source
                            noisy_slice_list = []
                            clean_slice_list = []

                            hop = int(
                                config.audio_size / config.moving_rate)  # This has to be an interger without any rounding
                            num_to_move = int((padded_song_shape - config.audio_size) / hop) + 1
                            for i in range(num_to_move):
                                start_point = int(i * hop)
                                end_point = int(start_point + config.audio_size)
                                noisy_slice = noisy_pad[start_point:end_point]  # ex: (1,16384)
                                noisy_slice_list.append(noisy_slice)
                                clean_slice = clean_pad[start_point:end_point]
                                clean_slice_list.append(clean_slice)

                            num_slice = len(noisy_slice_list)
                            noisy_stacked = np.array(noisy_slice_list)
                            clean_stacked = np.array(clean_slice_list)

                            # Separation
                            segments = []
                            spec_mix_list = []
                            spec_tgt_list = []
                            spec_est_list = []
                            spec_est1_list = []
                            spec_est2_list = []
                            alpha_list = []

                            for n, i in enumerate(range(num_slice)):
                                start_time = time.time()
                                print("processing {}: {}/{}".format(name, n, num_slice))
                                estimated_sample, spec_mix_, spec_tgt_, spec_est_, spec_est1_,  spec_est2_, alpha_ \
                                    = sess.run([self.graph_test.estimated, self.graph_test.input_spec,
                                                self.graph_test.source_spec, self.graph_test.masked_spec,
                                                self.graph_test.masked_spec1,  self.graph_test.masked_spec2,
                                                self.graph_test.alpha],
                                               feed_dict={self.mix_test: noisy_stacked[i:i + 1],
                                                          self.source_test: clean_stacked[i:i + 1]})
                                infer_time.append(time.time() - start_time)
                                masked_sample = np.expand_dims(mask, axis=0) * estimated_sample
                                segments.append(masked_sample)
                                alpha_list.append(np.mean(alpha_))

                                if (n+1) % 4 == 0:
                                    spec_mix_list.append(spec_mix_)
                                    spec_tgt_list.append(spec_tgt_)
                                    spec_est_list.append(spec_est_)
                                    spec_est1_list.append(spec_est1_)
                                    spec_est2_list.append(spec_est2_)

                            spec_mix_ = np.concatenate(spec_mix_list, axis=0)
                            spec_tgt_ = np.concatenate(spec_tgt_list, axis=0)
                            spec_est_ = np.concatenate(spec_est_list, axis=0)
                            spec_est1_ = np.concatenate(spec_est1_list, axis=0)
                            spec_est2_ = np.concatenate(spec_est2_list, axis=0)

                            # spectrograms
                            mag_mix_ = np.sqrt(np.square(spec_mix_[:, :, :, 0]) + np.square(spec_mix_[:, :, :, 1]))
                            mag_tgt_ = np.sqrt(np.square(spec_tgt_[:, :, :, 0]) + np.square(spec_tgt_[:, :, :, 1]))
                            mag_est_ = np.sqrt(np.square(spec_est_[:, :, :, 0]) + np.square(spec_est_[:, :, :, 1]))
                            mag_est1_ = np.sqrt(np.square(spec_est1_[:, :, :, 0]) + np.square(spec_est1_[:, :, :, 1]))
                            mag_est2_ = np.sqrt(np.square(spec_est2_[:, :, :, 0]) + np.square(spec_est2_[:, :, :, 1]))

                            mag_image1 = np.log10(np.concatenate((mag_mix_[0], mag_mix_[1], mag_mix_[2]), axis=1) + 1e-02)
                            mag_image2 = np.log10(np.concatenate((mag_tgt_[0], mag_tgt_[1], mag_tgt_[2]), axis=1) + 1e-02)
                            mag_image3 = np.log10(np.concatenate((mag_est_[0], mag_est_[1], mag_est_[2]), axis=1) + 1e-02)
                            mag_image4 = np.log10(np.concatenate((mag_est1_[0], mag_est1_[1], mag_est1_[2]), axis=1) + 1e-02)
                            mag_image5 = np.log10(np.concatenate((mag_est2_[0], mag_est1_[1], mag_est2_[2]), axis=1) + 1e-02)

                            fig = plt.figure(figsize=(25, 15))
                            fig.suptitle('Spectrograms', fontsize=20, family='serif')

                            font = {'family': 'serif',
                                    'color': 'darkred',
                                    'weight': 'normal',
                                    'size': 15,
                                    }

                            ax1 = fig.add_subplot(1, 5, 1)
                            ax1.set_title('Mixture', fontdict=font)
                            ax1.imshow(mag_image1, interpolation='nearest', aspect='auto', cmap='jet')
                            ax1.xaxis.set_tick_params(labelsize=15)
                            ax1.yaxis.set_tick_params(labelsize=15)
                            plt.gca().invert_yaxis()

                            ax2 = fig.add_subplot(1, 5, 2)
                            ax2.set_title('True Vocal', fontdict=font)
                            ax2.imshow(mag_image2, interpolation='nearest', aspect='auto', cmap='jet')
                            ax2.xaxis.set_tick_params(labelsize=15)
                            ax2.yaxis.set_tick_params(labelsize=15)
                            plt.gca().invert_yaxis()

                            ax3 = fig.add_subplot(1, 5, 3)
                            ax3.set_title('Estimated Vocal {:.3f}'.format(np.mean(alpha_)), fontdict=font)
                            ax3.imshow(mag_image3, interpolation='nearest', aspect='auto', cmap='jet')
                            ax3.xaxis.set_tick_params(labelsize=15)
                            ax3.yaxis.set_tick_params(labelsize=15)
                            plt.gca().invert_yaxis()

                            ax3 = fig.add_subplot(1, 5, 4)
                            ax3.set_title('Estimated-1 Vocal', fontdict=font)
                            ax3.imshow(mag_image4, interpolation='nearest', aspect='auto', cmap='jet')
                            ax3.xaxis.set_tick_params(labelsize=15)
                            ax3.yaxis.set_tick_params(labelsize=15)
                            plt.gca().invert_yaxis()

                            ax3 = fig.add_subplot(1, 5, 5)
                            ax3.set_title('Estimated-2 Vocal', fontdict=font)
                            ax3.imshow(mag_image5, interpolation='nearest', aspect='auto', cmap='jet')
                            ax3.xaxis.set_tick_params(labelsize=15)
                            ax3.yaxis.set_tick_params(labelsize=15)
                            plt.gca().invert_yaxis()

                            plt.tight_layout(pad=0.4, w_pad=1, h_pad=1.0)
                            plt.subplots_adjust(top=0.85)
                            fig.savefig("./Results/{}/{}_{}.png".format(folder, iter, os.path.splitext(name)[0]))
                            plt.close(fig)

                            # Post-processing(triangle mask)
                            # num_to_pad = int((config.audio_size/2)*(num_slice-1))
                            # temp = np.zeros(shape=(config.audio_size + num_to_pad))
                            temp = np.zeros(shape=(padded_song_shape))
                            for i in range(len(segments)):
                                start_point = int(i * (config.audio_size / config.moving_rate))
                                end_point = int(start_point + config.audio_size)
                                temp[start_point:end_point] = temp[start_point:end_point] + segments[i]

                            # Match the original song length
                            estimation = np.squeeze(temp[config.audio_size:config.audio_size + song_length])
                            # print("estimation shape: {}".format(estimation.shape))

                            # save separated source as audio
                            OUTPUT_FILENAME = "./Results/{}/{}_{}_{}_{}.wav".format(folder, iter, os.path.splitext(name)[0],
                                                                                  config.network, config.network2)
                            librosa.output.write_wav(OUTPUT_FILENAME, estimation, 16000)

                            s_target = clean * np.sum(estimation * clean) / np.sum(clean ** 2)
                            error = estimation - s_target
                            snr = 10 * np.log10(sum(s_target ** 2) / sum(error ** 2))
                            SNR_list.append(snr)

                            noise_type = os.path.splitext(name)[0].split("_")[2]
                            noise_snr = os.path.splitext(name)[0].split("_")[3]

                            ws.cell(row= 3, column= k % 3 + 2).value = noise_snr
                            ws.cell(row= k//3 + 4, column= k % 3 + 2).value = snr
                            ws.cell(row= k//3 + 4, column= 1).value = noise_type
                            ws.cell(row= 3, column= k % 3 + 6).value = noise_snr
                            ws.cell(row= k//3 + 4, column= k % 3 + 6).value = np.mean(np.array(alpha_list))

                        snr_mean = np.mean(np.array(SNR_list))
                        snr_std = np.std(np.array(SNR_list))
                        print("snr mean : {}, std : {}".format(snr_mean, snr_std))
                        ws.cell(row=1, column=2).value = "mean"
                        ws.cell(row=1, column=3).value = "std"
                        ws.cell(row=1, column=4).value = "infer_time per frame"
                        ws.cell(row=2, column=2).value = snr_mean
                        ws.cell(row=2, column=3).value = snr_std
                        ws.cell(row=2, column=4).value = np.mean(np.array(infer_time))
                        ws.cell(row=2, column=6).value = "alpha"

                        wb.save('./Results/{}/snr_{}_{}_{}.xlsx'.format(folder, iter, config.network, config.network2))
                        wb.close()

                    print("Save model for epoch {}".format(iter))
                    saver.save(sess, "./Results/CheckPoint/model.ckpt", global_step=(iter+1))
                    print("-------------------Model saved-------------------")

            coord.request_stop()
            coord.join(threads)

    def test(self, model_path, model_number):
        """
        Make Saving Directories
        """

        NUM_THREADS = 3
        configuration = tf.ConfigProto(inter_op_parallelism_threads=NUM_THREADS, \
                                       intra_op_parallelism_threads=NUM_THREADS, \
                                       allow_soft_placement=True, \
                                       device_count={'CPU': 3}, \
                                       )

        with tf.Session(config=configuration) as sess:
            path = os.path.join(os.getcwd(), 'Test', 'Model_{}'.format(config.model_number))
            os.makedirs(path, exist_ok=True)
            path = os.path.join(os.getcwd(), 'Test', 'Model_{}'.format(config.model_number), 'Audio')
            os.makedirs(path, exist_ok=True)

            saver = tf.train.Saver()
            saver.restore(sess, model_path+model_number)
            print("model is restored!!", model_path+model_number)

            # test songs
            moving_rate = config.moving_rate
            hop = int(config.audio_size / moving_rate)  # This has to be an interger without any rounding

            # mask = (signal.get_window('hanning', config.audio_size)) / (moving_rate / 2)
            mask = (1 / moving_rate)

            test_sample_path = glob(os.path.join(config.test_root, "*.wav"))
            print("{} audios for test".format(len(test_sample_path)))

            for k, song_path in enumerate(sorted(test_sample_path)):
                print("\n %d th song" % k)

                # Load song
                name = os.path.basename(song_path)
                print("Loading {}...".format(name))
                noisy, _ = librosa.core.load(song_path, sr=16000)

                print("song shape : ", noisy.shape)

                song_length = noisy.shape[0]
                num_frame = int(song_length / config.audio_size) + 1

                # pad zeros to song to make the length of it be the multiple of config.audio_size
                noisy_pad = np.lib.pad(noisy, (
                    config.audio_size, num_frame * config.audio_size - song_length + config.audio_size),
                                       'constant')
                padded_song_shape = noisy_pad.shape[0]

                # Slice mixture and source
                noisy_slice_list = []

                hop = int(
                    config.audio_size / config.moving_rate)  # This has to be an interger without any rounding
                num_to_move = int((padded_song_shape - config.audio_size) / hop) + 1
                for i in range(num_to_move):
                    start_point = int(i * hop)
                    end_point = int(start_point + config.audio_size)
                    noisy_slice = noisy_pad[start_point:end_point]  # ex: (1,16384)
                    noisy_slice_list.append(noisy_slice)

                num_slice = len(noisy_slice_list)
                noisy_stacked = np.array(noisy_slice_list)

                # Separation
                segments = []

                start_time = time.time()
                for n, i in enumerate(range(num_slice)):
                    print("processing {}: {}/{}".format(name, n, num_slice))
                    estimated_sample = sess.run(self.estimated,
                                   feed_dict={self.mix_test: noisy_stacked[i:i + 1]})
                    masked_sample = np.expand_dims(mask, axis=0) * estimated_sample
                    segments.append(masked_sample)
                print("processed time : {:.3f} ({:.3f} per 1s)".format(time.time()-start_time,(time.time()-start_time)/num_slice))

                # Post-processing(triangle mask)
                # num_to_pad = int((config.audio_size/2)*(num_slice-1))
                # temp = np.zeros(shape=(config.audio_size + num_to_pad))
                temp = np.zeros(shape=(padded_song_shape))
                for i in range(len(segments)):
                    start_point = int(i * (config.audio_size / config.moving_rate))
                    end_point = int(start_point + config.audio_size)
                    temp[start_point:end_point] = temp[start_point:end_point] + segments[i]

                # Match the original song length
                estimation = np.squeeze(temp[config.audio_size:config.audio_size + song_length])

                # save separated source as audio
                OUTPUT_FILENAME = os.path.join(path, "{}.wav".format(os.path.splitext(name)[0]))
                librosa.output.write_wav(OUTPUT_FILENAME, estimation, 16000)
