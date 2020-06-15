import matplotlib as mpl
mpl.use('Agg')
import matplotlib.pyplot as plt
import tensorflow as tf
from glob import glob
import os
from shutil import copyfile
from time import sleep, strftime, localtime
from Loader import Loader
from configuration import config, str2bool
from Utils import *
from layers import *
import librosa
import pickle
import openpyxl
import importlib


class FrontEnd(object):
    def __init__(self):
        pass

    def run(self, input, nfft=config.nfft, stride=config.stride, reuse=False, print_shape=False,
            is_training=False, fft_trainable=config.fft_trainable, hanning=config.hanning):
        self.input = input
        self.is_training = is_training

        self.output = front_end_cnn("stft",
                                    self.input,
                                    N=nfft,
                                    stride=stride,
                                    reuse=reuse,
                                    padding="SAME",
                                    print_shape=print_shape,
                                    dc_trainable=config.dc_trainable,
                                    fft_trainable=fft_trainable,
                                    hanning=hanning
                                    )
        return self.output


class MDPHDTest(object):
    def __init__(self, name):
        self.name = name
        self.FrontEnd = FrontEnd()

    def test(self, model_path, model_number):
        """
        Make Saving Directories
        """

        NUM_THREADS = 3
        configuration = tf.ConfigProto(inter_op_parallelism_threads=NUM_THREADS, \
                                       intra_op_parallelism_threads=NUM_THREADS, \
                                       allow_soft_placement=True, \
                                       device_count={'CPU': 3}, \
                                       )

        with tf.Session(config=configuration) as sess:
            meta_graph_path = './Result/train/CheckPoint/model.ckpt-{}.meta'.format(config.model_number)
            restorer = tf.train.import_meta_graph(meta_graph_path)
            graph = tf.get_default_graph()

            self.mix_test = graph.get_tensor_by_name("mdphd/mixture_sample:0")
            self.estimated = graph.get_tensor_by_name("mdphd/estimation_test:0")
            self.input_spec = graph.get_tensor_by_name("mdphd/input_spec_test:0")

            print("\nMODEL PATH : ", model_path)
            ckpt = tf.train.get_checkpoint_state(checkpoint_dir=model_path)
            ckpt_list = ckpt.all_model_checkpoint_paths

            model_exist = False
            for model in ckpt_list:
                # print(model.split("/")) # ['.', 'Result', '18-01-16-21-29-17', 'CheckPoint', 'model.ckpt-195000']
                if model_number in model.split("/")[-1]:
                    model_exist = True
                    print("Check point path : ", model)

                    """
                    Restore
                    """
                    restorer.restore(sess, model)
                    print('Restored!')

                    os.makedirs("./Result/test", exist_ok=True)
                    os.makedirs("./Result/test/EstimatedSampleSource".format(config.which_model, config.model_number),
                                exist_ok=True)  # make directory to save generated source audio by inference after training
                    continue
            if model_exist == False:
                print("No model number {}".format(config.model_number))

            moving_rate = config.moving_rate
            hop = int(config.audio_size / moving_rate)  # This has to be an interger without any rounding
            if config.use_mask == True:
                """Hanning Mask"""
                print("\nuse hanning_mask!")
                mask = (signal.get_window('hanning', config.audio_size)) / (moving_rate / 2)
            else:
                mask = (1 / moving_rate)

            # To do : make test set as numpy array and save
            test_sample_path = glob(os.path.join(config.test_root, "noisy/*.wav"))
            print("{} audios for test".format(len(test_sample_path)))

            SNR_list = []
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = 'song'
            ws.cell(row=1, column=2).value = 'SNR'

            for k, song_path in enumerate(test_sample_path):
                print("\n %d th song"%k)

                # Load song
                name = os.path.basename(song_path)
                print("Loading {}...".format(name))
                noisy, _ = librosa.core.load(song_path, sr=16000)
                clean, _ = librosa.core.load(os.path.join(config.test_root, "clean/{}".format(name)), sr=16000)

                print("song shape : ", noisy.shape)
                print("song shape : ", clean.shape)

                song_length = noisy.shape[0]
                num_frame = int(song_length/config.audio_size)+1

                # pad zeros to song to make the length of it be the multiple of config.audio_size
                noisy_pad = np.lib.pad(noisy, (config.audio_size,num_frame*config.audio_size-song_length+config.audio_size), config.pad_type)
                padded_song_shape = noisy_pad.shape[0]

                # Slice mixture and source
                noisy_slice_list = []
                source_slice_list = []

                num_to_move = int((padded_song_shape-config.audio_size)/hop) + 1
                for i in range(num_to_move): # it is already including extra segments(+1) since we padded the zeros to the last part of the song
                    start_point = int(i*hop)
                    end_point = int(start_point + config.audio_size)
                    noisy_slice = noisy_pad[start_point:end_point] # ex: (1,16384)
                    noisy_slice_list.append(noisy_slice)

                num_slice = len(noisy_slice_list)
                noisy_stacked = np.array(noisy_slice_list)

                # Separation
                segments = []
                is_scattered =0


                for n, i in enumerate(range(num_slice)):
                    print("processing {}: {}/{}".format(name, n, num_slice))
                    estimated_sample = sess.run(self.estimated, feed_dict={self.mix_test: noisy_stacked[i:i+1]})
                    masked_sample = np.expand_dims(mask, axis=0) * estimated_sample
                    segments.append(masked_sample)

                # Post-processing(triangle mask)
                # num_to_pad = int((config.audio_size/2)*(num_slice-1))
                # temp = np.zeros(shape=(config.audio_size + num_to_pad))
                temp = np.zeros(shape=(padded_song_shape))
                for i in range(len(segments)):
                    start_point = int(i*(config.audio_size/moving_rate))
                    end_point = int(start_point + config.audio_size)
                    temp[start_point:end_point] = temp[start_point:end_point] + segments[i]

                # Match the original song length
                estimation = np.squeeze(temp[config.audio_size:config.audio_size+song_length])
                # print("estimation shape: {}".format(estimation.shape))

                if k<=5:
                    # save separated source as audio
                    OUTPUT_FILENAME1 = "./Result/test/EstimatedSampleSource/{}".format(name)
                    OUTPUT_FILENAME2 = "./Result/test/EstimatedSampleSource/clean_{}".format(name)
                    OUTPUT_FILENAME3 = "./Result/test/EstimatedSampleSource/noisy_{}".format(name)
                    librosa.output.write_wav(OUTPUT_FILENAME1, estimation, 16000)
                    librosa.output.write_wav(OUTPUT_FILENAME2, clean, 16000)
                    librosa.output.write_wav(OUTPUT_FILENAME3, noisy, 16000)

                s_target = clean * np.sum(estimation * clean) / np.sum(clean**2)
                error = estimation - s_target
                snr = 10 * np.log10(sum(s_target ** 2) / sum(error ** 2))

                SNR_list.append(snr)

                ws.cell(row=k+4, column=1).value = name
                ws.cell(row=k+4, column=2).value = snr

                if k ==10:
                    break

            snr_mean = np.mean(np.array(SNR_list))
            snr_std = np.std(np.array(SNR_list))
            print("snr mean : {}, std : {}".format(snr_mean, snr_std))
            ws.cell(row=1, column=4).value = "mean"
            ws.cell(row=1, column=5).value = "std"
            ws.cell(row=2, column=4).value = snr_mean
            ws.cell(row=2, column=5).value = snr_std

            wb.save('./Result/test/snr_{}.xlsx'.format(config.model_number))
            wb.close()

            print("Test Finished!")


if __name__=="__main__":
    network=MDPHDTest("mdphd")
    network.test(model_path="./Result/train/CheckPoint/", model_number="model.ckpt-{}".format(config.model_number))
